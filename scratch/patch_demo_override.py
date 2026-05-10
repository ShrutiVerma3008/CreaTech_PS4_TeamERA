"""
scratch/patch_demo_override.py
Adds floor_override column to data/demo_tower_40floors.xlsx.
All 40 rows: False. F36-F40: True (intentional architectural exceptions).
"""
import openpyxl
import os

DEMO_PATH = "data/demo_tower_40floors.xlsx"

wb = openpyxl.load_workbook(DEMO_PATH)
ws = wb.active

# Read header row
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
print("Current headers:", headers)

# Add floor_override header if not already present
if "floor_override" not in headers:
    new_col_idx = len(headers) + 1  # 1-indexed
    ws.cell(row=1, column=new_col_idx, value="floor_override")
    print(f"Added 'floor_override' header at column {new_col_idx}")
else:
    new_col_idx = headers.index("floor_override") + 1
    print(f"'floor_override' already exists at column {new_col_idx}")

# Find floor_id column index
floor_id_col_idx = headers.index("floor_id") + 1  # 1-indexed

# Set values: False for all, True for F36-F40
OVERRIDE_FLOORS = {"F36", "F37", "F38", "F39", "F40"}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    fid = row[floor_id_col_idx - 1].value
    override_val = fid in OVERRIDE_FLOORS
    ws.cell(row=row[0].row, column=new_col_idx, value=override_val)
    if override_val:
        print(f"  Row {row[0].row}: floor_id={fid} -> floor_override=True")

wb.save(DEMO_PATH)
print(f"\nSaved {DEMO_PATH}")

# Verify
wb2 = openpyxl.load_workbook(DEMO_PATH)
ws2 = wb2.active
headers2 = [cell.value for cell in next(ws2.iter_rows(min_row=1, max_row=1))]
assert "floor_override" in headers2, "FAIL: floor_override column missing after save"
override_true_count = sum(
    1 for row in ws2.iter_rows(min_row=2, values_only=True)
    if row[headers2.index("floor_override")] == True
)
print(f"Verified: {override_true_count} rows have floor_override=True (expected 5)")
assert override_true_count == 5, f"FAIL: expected 5 True rows, got {override_true_count}"
print("Patch complete — demo file verified.")
